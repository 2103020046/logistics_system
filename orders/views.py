# views.py
import json

from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Order, Item
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import logging
from django.shortcuts import get_object_or_404, redirect
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.http import HttpResponse
import openpyxl
from datetime import datetime

# 配置日志记录
logger = logging.getLogger(__name__)


@csrf_exempt
def create_order(request):
    if request.method == 'POST':
        try:
            # 解析表单数据
            # 在create_order视图的order_data中添加
            order_data = {
                'user': request.user,
                'company_name': request.POST.get('companyName'),  # 新增公司名称字段
                'order_number': request.POST.get('orderNo'),  # 运单号
                'sender': request.POST.get('senderName', ''),  # 发货人
                'sender_phone': request.POST.get('senderPhone', ''),  # 发货人手机号
                'sender_address': request.POST.get('senderAddress', ''),  # 发货详细地址
                'product_code': request.POST.get('productCode', ''),  # 货号
                'receiver': request.POST.get('receiverName', ''),  # 收货方
                'receiver_phone': request.POST.get('receiverPhone', ''),  # 收货人手机号
                'receiver_address': request.POST.get('receiverAddress', ''),  # 收货详细地址
                'total_freight': request.POST.get('totalFee', 0),  # 总费用
                'payment_method': request.POST.get('paymentMethod', ''),  # 支付方式
                'return_requirement': request.POST.get('returnRequirement', ''),  # 回单要求
                'other_expenses': request.POST.get('otherExpenses', 0),  # 其他支出
                'expense_details': request.POST.get('feeDescription', ''),  # 费用说明
                'carrier': request.POST.get('carrier', ''),  # 承运商
                'carrier_address': request.POST.get('carrierAddress', ''),  # 到站地址
                'arrival_address': request.POST.get('arrivalAddress', ''),  # 发站地址
                'departure_station_phone': request.POST.get('departureStationPhone', None),  # 发站查询电话
                'arrival_station_phone': request.POST.get('arrivalStationPhone', ''),  # 到站查询电话
                'customer_order_no': request.POST.get('customerOrderNo', ''),  # 客户单号
                'date': request.POST.get('date'),  # 日期
                'departure_station': request.POST.get('departureStation'),  # 发站
                'arrival_station': request.POST.get('arrivalStation'),  # 到站
                'transport_method': request.POST.get('transportMethod'),  # 运输方式
                'delivery_method': request.POST.get('deliveryMethod'),  # 交货方式
                'sender_sign': request.POST.get('senderSign'),  # 发货人签名
                'receiver_sign': request.POST.get('receiverSign'),  # 收货人签名
                'id_card': request.POST.get('idCard'),  # 身份证号
                'order_maker': request.POST.get('orderMaker'),  # 制单人
                'fee_wan': request.POST.get('fee_wan', ''),
                'fee_qian': request.POST.get('fee_qian', ''),
                'fee_bai': request.POST.get('fee_bai', ''),
                'fee_shi': request.POST.get('fee_shi', ''),
                'fee_ge': request.POST.get('fee_ge', ''),
            }

            # 确保所有必需字段都有值
            required_fields = ['order_number', 'sender', 'receiver', 'receiver_phone', 'receiver_address']
            for field in required_fields:
                if not order_data[field]:
                    return JsonResponse({'status': 'error', 'message': f'Missing required field: {field}'}, status=400)

            # 创建Order实例
            order = Order.objects.create(**{k: v for k, v in order_data.items() if v is not None})

            # 处理商品数据
            items = []
            i = 0
            while True:
                item_key = f'items[{i}][productName]'
                if item_key not in request.POST:  # 检查是否存在该键
                    break
                item_data = {
                    'order_id': order.id,
                    'item_name': request.POST[f'items[{i}][productName]'],  # 品名
                    'package_type': request.POST[f'items[{i}][packageType]'],  # 包装
                    'quantity': int(request.POST[f'items[{i}][quantity]']),  # 件数
                    'weight': float(request.POST[f'items[{i}][weight]']),  # 重量
                    'volume': float(request.POST[f'items[{i}][volume]']),  # 体积
                    'delivery_charge': float(request.POST.get(f'items[{i}][deliveryCharge]', 0)),  # 送（提）货费
                    'insurance_fee': float(request.POST.get(f'items[{i}][insuranceFee]', 0)),  # 保险费
                    'packaging_fee': float(request.POST.get(f'items[{i}][packagingFee]', 0)),  # 包装费
                    'goods_value': float(request.POST.get(f'items[{i}][goodsValue]', 0)),  # 货物价值
                    'remarks': request.POST.get(f'items[{i}][remarks]', ''),  # 备注
                    'freight': float(request.POST[f'items[{i}][freight]'])  # 运费
                }
                items.append(Item.objects.create(**item_data))
                i += 1

            # 返回包含 orderId 的 JSON 响应
            return JsonResponse({
                'status': 'success',
                'message': 'Order created successfully',
                'orderId': order.id  # 添加 orderId 到响应中
            })

        except Exception as e:
            logger.error(f"Unexpected error creating order: {str(e)}", exc_info=True)
            return JsonResponse(
                {'status': 'error', 'message': f'{str(e)}, An unexpected error occurred. Please try again later.'},
                status=500)


def index(request):
    return render(request, 'index.html')


def custom_login_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return view_func(request, *args, **kwargs)

    return wrapper


@custom_login_required
@login_required
def orders(request):
    return render(request, 'order.html')


@custom_login_required
@login_required
def order_history(request):
    # 获取查询参数
    receiver = request.GET.get('receiver', '')
    order_number = request.GET.get('order_number', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # 构建查询条件
    queryset = Order.objects.filter(user=request.user)
    
    if receiver:
        queryset = queryset.filter(receiver__icontains=receiver)
    if order_number:
        queryset = queryset.filter(order_number=order_number)
    if start_date and end_date:
        queryset = queryset.filter(date__gte=start_date, date__lte=end_date)
    
    # 按日期倒序排序
    queryset = queryset.order_by('-date')

    # 分页设置
    paginator = Paginator(queryset, 10)
    page = request.GET.get('page')

    try:
        orders_page = paginator.page(page)
    except PageNotAnInteger:
        orders_page = paginator.page(1)
    except EmptyPage:
        orders_page = paginator.page(paginator.num_pages)

    # 保留搜索参数在分页链接中
    params = request.GET.copy()
    if 'page' in params:
        del params['page']
    orders_page.query_params = params.urlencode()

    return render(request, 'order_history.html', {
        'orders': orders_page,
        'search_params': {
            'receiver': receiver,
            'order_number': order_number,
            'start_date': start_date,
            'end_date': end_date
        }
    })


@csrf_exempt
@login_required
def get_order_detail(request, order_id):
    try:
        # 获取当前用户的订单
        order = Order.objects.get(id=order_id, user=request.user)
        items = order.items.all()  # 获取所有关联的商品信息
        items_data = [{
            'item_name': item.item_name,
            'package_type': item.package_type,
            'quantity': item.quantity,
            'weight': str(item.weight),  # 转换为字符串以避免 JSON 序列化问题
            'volume': str(item.volume),
            'delivery_charge': str(item.delivery_charge),
            'insurance_fee': str(item.insurance_fee),
            'packaging_fee': str(item.packaging_fee),
            'goods_value': str(item.goods_value),
            'remarks': str(item.remarks),
            'freight': str(item.freight),
        } for item in items]

        return JsonResponse({
            'order_number': order.order_number,
            'sender': order.sender,
            'sender_phone': order.sender_phone,
            'sender_address': order.sender_address,
            'product_code': order.product_code,
            'receiver': order.receiver,
            'receiver_phone': order.receiver_phone,
            'receiver_address': order.receiver_address,
            'total_freight': str(order.total_freight),  # 转换为字符串
            'payment_method': order.payment_method,
            'other_expenses': str(order.other_expenses),  # 转换为字符串
            'expense_details': order.expense_details,
            'carrier': order.carrier,
            'carrier_address': order.carrier_address,
            'arrival_address': order.arrival_address,
            'return_requirement': order.return_requirement,
            'departure_station_phone': order.departure_station_phone,
            'arrival_station_phone': order.arrival_station_phone,
            'customer_order_no': order.customer_order_no,
            'date': order.date,
            'departure_station': order.departure_station,
            'arrival_station': order.arrival_station,
            'transport_method': order.transport_method,
            'delivery_method': order.delivery_method,
            'sender_sign': order.sender_sign,
            'receiver_sign': order.receiver_sign,
            'id_card': order.id_card,
            'order_maker': order.order_maker,
            'items': items_data,
        })
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)


@login_required
def edit_order(request, order_id):
    order = get_object_or_404(Order.objects.prefetch_related('items'), id=order_id, user=request.user)
    context = {
        'order': order,
    }
    if request.method == 'POST':
        # 更新订单基本信息
        order.sender = request.POST.get('senderName')
        order.receiver = request.POST.get('receiverName')
        order.sender_phone = request.POST.get('senderPhone')
        order.receiver_phone = request.POST.get('receiverPhone')
        order.sender_address = request.POST.get('senderAddress')
        order.receiver_address = request.POST.get('receiverAddress')
        order.product_code = request.POST.get('productCode')
        order.total_freight = request.POST.get('totalFee')
        order.payment_method = request.POST.get('paymentMethod')
        order.return_requirement = request.POST.get('returnRequirement')
        order.other_expenses = request.POST.get('otherExpenses')
        order.expense_details = request.POST.get('feeDescription')
        order.carrier = request.POST.get('carrier')
        order.carrier_branch = request.POST.get('carrierBranch')
        order.departure_station_phone = request.POST.get('departureStationPhone')
        order.arrival_station = request.POST.get('arrivalStation')
        order.arrival_station_phone = request.POST.get('arrivalStationPhone')
        order.transit_fee = request.POST.get('transitFee')
        order.date = request.POST.get('date')
        order.departure_station = request.POST.get('departureStation')
        order.transport_method = request.POST.get('transportMethod')
        order.delivery_method = request.POST.get('deliveryMethod')
        order.sender_sign = request.POST.get('senderSign')
        order.receiver_sign = request.POST.get('receiverSign')
        order.id_card = request.POST.get('idCard')
        order.order_maker = request.POST.get('orderMaker')
        order.save()

        # 更新商品项
        order.items.all().delete()  # 删除旧商品项
        index = 0
        while True:
            product_name = request.POST.get(f'items[{index}][productName]')
            if not product_name:
                break
            Item.objects.create(
                order=order,
                item_name=product_name,
                package_type=request.POST.get(f'items[{index}][packageType]'),
                quantity=request.POST.get(f'items[{index}][quantity]'),
                weight=request.POST.get(f'items[{index}][weight]'),
                volume=request.POST.get(f'items[{index}][volume]'),
                delivery_charge=request.POST.get(f'items[{index}][deliveryCharge]'),
                insurance_fee=request.POST.get(f'items[{index}][insuranceFee]'),
                packaging_fee=request.POST.get(f'items[{index}][packagingFee]'),
                goods_value=request.POST.get(f'items[{index}][goodsValue]'),
                remarks=request.POST.get(f'items[{index}][remarks]'),
                freight=request.POST.get(f'items[{index}][freight]'),
            )
            index += 1
        return redirect('order_history')

    return render(request, 'edit_order.html', {'order': order})


def delete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order.delete()
    return JsonResponse({'status': 'success'})


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('/')
        else:
            messages.error(request, '用户名或密码错误，请重试。')
    return render(request, 'login.html')


def register_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        if password1 != password2:
            messages.error(request, '两次输入的密码不一致，请重试。')
        else:
            user = User.objects.create_user(username=username, password=password1)
            login(request, user)  # 自动登录新用户
            messages.success(request, '注册成功！')
            return redirect('/')
    return render(request, 'register.html')


def logout_view(request):
    logout(request)
    return redirect('/')


@csrf_exempt
def update_order(request, order_id):
    if request.method == 'PUT':
        try:
            # 获取订单对象
            order = Order.objects.get(id=order_id)

            # 解析请求体中的JSON数据
            data = json.loads(request.body.decode('utf-8'))

            # 更新订单信息
            for item in data.get('items', []):
                # 假设你有一个模型 Item 来保存每个货物的信息
                item_obj, created = Item.objects.update_or_create(
                    order=order,
                    index=item['index'],
                    defaults={
                        'product_name': item['productName'],
                        'package_type': item['packageType'],
                        'quantity': item['quantity'],
                        'weight': item['weight'],
                        'volume': item['volume'],
                        'delivery_charge': item['deliveryCharge'],
                        'insurance_fee': item['insuranceFee'],
                        'packaging_fee': item['packagingFee'],
                        'goods_value': item['goodsValue'],
                        'freight': item['freight'],
                        'remarks': item['remarks']
                    }
                )

            return JsonResponse({'status': 'success', 'message': '订单更新成功'}, status=200)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': '只接受PUT请求'}, status=405)


def export_orders(request):
    if request.method == 'POST':
        try:
            order_ids = json.loads(request.POST.get('order_ids', '[]'))
            orders = Order.objects.filter(id__in=order_ids).prefetch_related('items').order_by('date')
            
            # 获取第一个订单日期（如果存在）并转换为datetime对象
            if orders and orders[0].date:
                first_date = datetime.strptime(orders[0].date, "%Y-%m-%d")
                filename_date = first_date.strftime("%m月%d日")
            else:
                first_date = datetime.now()
                filename_date = first_date.strftime("%m月%d日")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 设置标题（使用第一个订单日期）
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = f"{filename_date}发货明细"
            title_cell.font = openpyxl.styles.Font(bold=True, size=14)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            # 设置表头（新增短途列）
            headers = ['日期', '客户名', '品名', '包装', '件数', '重量', '方数', 
                     '打包费', '短途', '运费', '备注', '地址']
            ws.append(headers)
            
            # 设置列宽（地址列加宽）
            ws.column_dimensions['L'].width = 30
            ws.column_dimensions['B'].width = 25

            # 初始化合计变量
            total_quantity = 0
            total_weight = 0
            total_volume = 0
            total_freight = 0

            # 填充数据
            for order in orders:
                order_date = datetime.strptime(order.date, "%Y-%m-%d") if order.date else datetime.now()
                
                for item in order.items.all():
                    row = [
                        order_date.strftime("%m月%d日"),
                        order.receiver,
                        item.item_name,
                        item.package_type,
                        item.quantity,
                        item.weight,
                        item.volume,
                        item.packaging_fee,
                        "",  # 短途列留空
                        item.freight,
                        order.payment_method,
                        order.receiver_address
                    ]
                    ws.append(row)
                    
                    # 累加合计值
                    total_quantity += item.quantity
                    total_weight += item.weight
                    total_volume += item.volume
                    total_freight += item.freight

            # 添加合计行
            if orders:  # 只有当有订单时才添加合计行
                ws.append([
                    "合计", "", "", "", 
                    total_quantity, 
                    total_weight, 
                    total_volume, 
                    "", "", 
                    total_freight, 
                    "", ""
                ])
            
            # 设置单元格居中
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                headers={'Content-Disposition': f'attachment; filename="{filename_date}发货明细.xlsx"'}
            )
            wb.save(response)
            return response
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)



# 获取当天订单数量，生成运单号
@csrf_exempt
def get_today_order_count(request):
    if request.method == 'GET':
        today = datetime.now().date()
        count = Order.objects.filter(date=today.strftime("%Y-%m-%d")).count()
        return JsonResponse({'count': count})
    return JsonResponse({'count': 0})